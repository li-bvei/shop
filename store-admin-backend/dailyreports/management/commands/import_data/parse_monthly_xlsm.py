"""Regenerate daily_reports_2026.json from the monthly 売上日報 spreadsheets.

Not wired into Django — a standalone helper kept next to the data it produces
(both are gitignored: real business data). `import_daily_reports_2026` reads
the JSON this writes and does the idempotent DB upsert.

Usage:
    python parse_monthly_xlsm.py "8月 (1).xlsm:8" "9月.xlsm:9"

Each arg is  <filename in ~/Downloads>:<month number>.  Months given here are
re-parsed in full and replace whatever the current JSON holds for that month;
every other month in the JSON is kept verbatim. Run with no args to just
re-report the current JSON.

The sheet layout drifts between days (a payment row is reused for whatever
label that day happened to use), so everything is found by the label text in
column A / D, never by fixed coordinates.
"""
import json
import sys
from collections import Counter
from pathlib import Path

import openpyxl

YEAR = 2026
DOWNLOADS = Path.home() / 'Downloads'
JSON_PATH = Path(__file__).parent / 'daily_reports_2026.json'

# 現金 is deliberately excluded — DailyReport.save() recomputes it as
# total_revenue minus every other method. Any label not in this map is
# dropped and its amount folds back into that recomputed cash figure
# (total_revenue stays exact): 昨日欠 / M/D欠 are cash-shortfall carry-overs,
# not tenders; 商品卷 / GOTO / 食事券 are voucher types with no matching
# PaymentMethodDef on 心斎橋 anymore. The original 2026 import did the same
# (it only ever looked for the labels it knew).
PAYMENT_LABEL_TO_CODE = {
    'カード': 'creditCard',
    '電子マネー': 'emoney',
    '交通系': 'emoney',            # transit IC cards — same 電子マネー bucket
    'プレミアム': 'osakaCoupon',    # 大阪プレミアム(付)商品券 — a default method
    'ぐるなび金券': 'gurunaviCoupon',
    'HP金券': 'hpCoupon',
    '食事券': 'mealVoucher',
    '売掛金': 'onCredit',
    '微信': 'wechat',
    'paypay': 'paypay',
    'プリペイドカード': 'prepaidCard',
    'ポイント': 'points',
}


def build_label_index(ws, col):
    index = {}
    for row in ws.iter_rows(min_col=col, max_col=col):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip():
                index.setdefault(cell.value.strip(), cell.row)
    return index


def num(v):
    return v if isinstance(v, (int, float)) else 0


def parse_day_sheet(ws, date_str, warnings, dropped):
    a = build_label_index(ws, 1)
    d = build_label_index(ws, 4)

    revenue_row = a.get('売上￥')
    if revenue_row is None:
        warnings.append(f'{date_str}: no 売上￥ row, skipped')
        return None
    if ws.cell(row=revenue_row, column=2).value is None:
        return None  # blank template day, not an error

    result = {
        'date': date_str,
        'total_revenue': num(ws.cell(row=revenue_row, column=2).value),
        'total_customers': num(ws.cell(row=revenue_row, column=4).value),
        'group_count': num(ws.cell(row=revenue_row, column=6).value),
        'morning_revenue': 0,
        'morning_customers': 0,
        'morning_group_count': 0,
        'person_in_charge_raw': '',
        'payment_amounts': {},
        'expenses': [],
    }

    person_row = a.get('記入者')
    if person_row and ws.cell(row=person_row, column=2).value:
        result['person_in_charge_raw'] = str(ws.cell(row=person_row, column=2).value).strip()

    lunch_row = a.get('ランチ')
    if lunch_row:
        result['morning_revenue'] = num(ws.cell(row=lunch_row, column=2).value)
        result['morning_customers'] = num(ws.cell(row=lunch_row, column=4).value)
        result['morning_group_count'] = num(ws.cell(row=lunch_row, column=6).value)
    else:
        warnings.append(f'{date_str}: no ランチ row')

    start = a.get('売上金種')
    end = a.get('現金外合計')
    if start and end:
        for r in range(start + 1, end):
            label = ws.cell(row=r, column=1).value
            if not (isinstance(label, str) and label.strip()):
                continue
            lab = label.strip()
            if lab == '現金':
                continue
            val = ws.cell(row=r, column=2).value
            if val in (None, 0):
                continue
            code = PAYMENT_LABEL_TO_CODE.get(lab)
            if code:
                result['payment_amounts'][code] = result['payment_amounts'].get(code, 0) + num(val)
            else:
                dropped.append((date_str, lab, num(val)))

    header_row = d.get('仕入れ業者')
    if header_row and end and end > header_row:
        for r in range(header_row + 1, end):
            vendor = ws.cell(row=r, column=4).value
            item = ws.cell(row=r, column=6).value
            amount = ws.cell(row=r, column=8).value
            if vendor or item or amount:
                result['expenses'].append({
                    'itemName': str(vendor).strip() if vendor else '',
                    'purpose': str(item).strip() if item else '',
                    'amount': num(amount),
                })
    elif not header_row:
        warnings.append(f'{date_str}: no 仕入れ業者 header')

    return result


def main(args):
    existing = json.loads(JSON_PATH.read_text(encoding='utf-8')) if JSON_PATH.exists() else []

    warnings, dropped, new_records = [], [], []
    for arg in args:
        fname, month = arg.rsplit(':', 1)
        month = int(month)
        wb = openpyxl.load_workbook(DOWNLOADS / fname, data_only=True, keep_vba=False)
        for sn in wb.sheetnames:
            if not sn.endswith('日'):
                continue
            date_str = f'{YEAR}-{month:02d}-{int(sn[:-1]):02d}'
            rec = parse_day_sheet(wb[sn], date_str, warnings, dropped)
            if rec is not None:
                new_records.append(rec)

    replaced_months = {r['date'][:7] for r in new_records}
    merged = [r for r in existing if r['date'][:7] not in replaced_months] + new_records
    merged.sort(key=lambda r: r['date'])

    if args:
        JSON_PATH.write_text(json.dumps(merged, ensure_ascii=False, indent=1), encoding='utf-8')

    print(f'{JSON_PATH.name}: {len(merged)} records  {merged[0]["date"]}..{merged[-1]["date"]}')
    print('  by month:', dict(Counter(r['date'][:7] for r in merged)))
    if args:
        print(f'  replaced months: {sorted(replaced_months)}  ({len(new_records)} day records)')
        persons, codes = Counter(), Counter()
        for r in new_records:
            persons.update(r['person_in_charge_raw'].replace('　', ' ').split())
            codes.update(r['payment_amounts'].keys())
        print('  persons (day counts):', dict(persons))
        print('  payment codes (day counts):', dict(codes))
        if dropped:
            print(f'  DROPPED payment labels ({len(dropped)}):')
            for dt, lab, amt in dropped:
                print(f'    {dt}  {lab!r}  {amt:,}')
        if warnings:
            print(f'  WARNINGS ({len(warnings)}):')
            for w in warnings:
                print('   ', w)


if __name__ == '__main__':
    main(sys.argv[1:])
